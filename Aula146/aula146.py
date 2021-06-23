"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform

# pedidos = openpyxl.load_workbook('pedidos.xlsx')
# nome_planilhas = pedidos.sheetnames
# planilhas1 = pedidos['Página1']
#
# for linha in range(5, 16):
#     numero_pedido = linha - 1
#     planilhas1.cell(linha, 1).value = numero_pedido
#     planilhas1.cell(linha, 2).value = 1200 + linha
#
#     preco = round(uniform(10, 100), 2)
#     planilhas1.cell(linha, 3).value = preco
#
# pedidos.save('novo_planilha.xlsx')

# for campo in planilhas1['b']:
#     print(campo.value)

# for linha in planilhas1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

# for linha in planilhas1:
#     if linha[0].value is not None:
#         print(linha[0].value, end=" ")
#     if linha[1].value is not None:
#         print(linha[1].value, end=" ")
#     if linha[2].value is not None:
#         print(linha[2].value)

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Caio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Maria {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'João {linha} {round(uniform(10, 100), 2)}'
    # numero_pedido = linha - 1
    # planilhas2.cell(linha, 1).value = numero_pedido
    # planilhas2.cell(linha, 2).value = 1200 + linha
    #
    # preco = round(uniform(10, 100), 2)
    # planilhas2.cell(linha, 3).value = preco

planilha.save('novo_planilha.xlsx')
